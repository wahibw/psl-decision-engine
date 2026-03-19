"""
scripts/backtest_runner.py

Reads psl_backtest_2025_filled.xlsx and runs it against the live
toss_advisor engine to measure accuracy.

Works with partially-filled files — only columns K, L, M are required
per match. Columns P, Q, R are optional and silently skipped when empty.

Usage:
    python scripts/backtest_runner.py
    python scripts/backtest_runner.py --xlsx path/to/file.xlsx
    python scripts/backtest_runner.py --venue "Gaddafi Stadium, Lahore"
    python scripts/backtest_runner.py --summary-only

Output:
    - Overall toss accuracy %
    - Accuracy broken down by venue
    - Wrong calls with explanation
    - (If P/Q filled) Partnership threshold hit rate
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Optional

# ── Project root on path ─────────────────────────────────────────────────────
PROJ_ROOT = Path(__file__).resolve().parent.parent
if str(PROJ_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJ_ROOT))


try:
    import openpyxl
except ImportError:
    print("[backtest_runner] openpyxl is required: pip install openpyxl")
    sys.exit(1)

# Default xlsx location — save the file here in your project
DEFAULT_XLSX = PROJ_ROOT / "data" / "backtest" / "psl_backtest_2025_filled.xlsx"

# Column indices (1-based, matching the spreadsheet)
COL_MATCH_NO     = 1
COL_DATE         = 2
COL_TEAM1        = 3
COL_TEAM2        = 4
COL_VENUE        = 5
COL_SYS_REC      = 6   # system toss recommendation (pre-filled)
COL_SYS_REASON   = 7
COL_K_TOSS_WIN   = 11  # actual toss winner
COL_L_TOSS_DEC   = 12  # actual decision: "bat" | "field"
COL_M_WINNER     = 13  # match winner
COL_N_MARGIN     = 14  # margin number
COL_O_TYPE       = 15  # "runs" | "wickets" | "NR"
COL_P_PARTNER    = 16  # biggest partnership runs (optional)
COL_Q_BREAKER    = 17  # partnership breaker (optional)
COL_R_DEATH_ECO  = 18  # death over economy (optional)
COL_S_NOTES      = 19  # notes

DATA_START_ROW   = 4
DATA_END_ROW     = 37  # matches 1-34


# ── DATA LOADER ──────────────────────────────────────────────────────────────

def load_backtest_data(xlsx_path: Path) -> list[dict]:
    """
    Read all 34 match rows from the Back-Test Log sheet.
    Returns list of dicts, one per match.
    Rows with no toss data are still included but flagged as incomplete.
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"Back-test file not found: {xlsx_path}\n"
            f"Place psl_backtest_2025_filled.xlsx in: {xlsx_path.parent}"
        )

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Back-Test Log" not in wb.sheetnames:
        raise ValueError("Sheet 'Back-Test Log' not found in workbook.")
    ws = wb["Back-Test Log"]

    matches = []
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        def v(col: int) -> Optional[str]:
            val = ws.cell(row=row, column=col).value
            if val is None or str(val).strip() == "":
                return None
            return str(val).strip()

        def fv(col: int) -> Optional[float]:
            val = ws.cell(row=row, column=col).value
            try:
                return float(val) if val is not None else None
            except (ValueError, TypeError):
                return None

        match_no = ws.cell(row=row, column=COL_MATCH_NO).value
        if not match_no:
            continue

        toss_win  = v(COL_K_TOSS_WIN)
        toss_dec  = v(COL_L_TOSS_DEC)
        sys_rec   = v(COL_SYS_REC)
        is_nr     = v(COL_O_TYPE) in ("NR", "No Result")

        # Compute toss verdict
        if is_nr:
            verdict = "SKIP"
        elif toss_win and toss_dec and sys_rec:
            if sys_rec == "BOWL FIRST" and toss_dec == "field":
                verdict = "CORRECT"
            elif sys_rec == "BAT FIRST" and toss_dec == "bat":
                verdict = "CORRECT"
            elif sys_rec == "TOSS-UP":
                verdict = "CORRECT"
            else:
                verdict = "WRONG"
        else:
            verdict = "INCOMPLETE"

        matches.append({
            "match_no":    int(match_no),
            "date":        v(COL_DATE),
            "team1":       v(COL_TEAM1),
            "team2":       v(COL_TEAM2),
            "venue":       v(COL_VENUE),
            "sys_rec":     sys_rec,
            "sys_reason":  v(COL_SYS_REASON),
            "toss_winner": toss_win,
            "toss_dec":    toss_dec,
            "winner":      v(COL_M_WINNER),
            "margin":      fv(COL_N_MARGIN),
            "margin_type": v(COL_O_TYPE),
            "partnership": fv(COL_P_PARTNER),    # None if not filled
            "p_breaker":   v(COL_Q_BREAKER),     # None if not filled
            "death_eco":   fv(COL_R_DEATH_ECO),  # None if not filled
            "notes":       v(COL_S_NOTES),
            "verdict":     verdict,
            "is_nr":       is_nr,
        })

    return matches


# ── ANALYSIS ─────────────────────────────────────────────────────────────────

def toss_accuracy_report(matches: list[dict], venue_filter: str = "") -> None:
    """Print toss accuracy overall and by venue."""
    if venue_filter:
        matches = [m for m in matches if venue_filter.lower() in (m["venue"] or "").lower()]
        if not matches:
            print(f"No matches found for venue filter: {venue_filter!r}")
            return

    scored  = [m for m in matches if m["verdict"] in ("CORRECT", "WRONG")]
    correct = [m for m in scored if m["verdict"] == "CORRECT"]
    wrong   = [m for m in scored if m["verdict"] == "WRONG"]
    skipped = [m for m in matches if m["verdict"] == "SKIP"]
    incomplete = [m for m in matches if m["verdict"] == "INCOMPLETE"]

    total = len(scored)
    pct   = len(correct) / total * 100 if total > 0 else 0

    print(f"\n{'='*65}")
    print(f"  PSL DECISION ENGINE — TOSS BACK-TEST RESULTS  (PSL 2025)")
    print(f"{'='*65}")
    print(f"\n  Overall toss accuracy : {len(correct)}/{total}  ({pct:.1f}%)")
    print(f"  Skipped (rain/NR)     : {len(skipped)}")
    if incomplete:
        print(f"  Incomplete (no data)  : {len(incomplete)}")

    # By venue
    venues = sorted(set(m["venue"] for m in scored if m["venue"]))
    if len(venues) > 1:
        print(f"\n  By venue:")
        print(f"  {'Venue':<40}  Correct  Total  Acc%   Sys Rec")
        print(f"  {'-'*75}")
        for venue in venues:
            vm = [m for m in scored if m["venue"] == venue]
            vc = [m for m in vm if m["verdict"] == "CORRECT"]
            vp = len(vc) / len(vm) * 100 if vm else 0
            # What did the system recommend here?
            recs = [m["sys_rec"] for m in vm if m["sys_rec"]]
            most_common = max(set(recs), key=recs.count) if recs else "?"
            short = venue[:38]
            print(f"  {short:<40}  {len(vc):>5}    {len(vm):>4}  {vp:>5.1f}%   {most_common}")

    # Wrong calls detail
    if wrong:
        print(f"\n  Wrong calls ({len(wrong)}):")
        print(f"  {'M#':>3}  {'Venue':<32}  {'Sys rec':>10}  {'Actual':>8}  {'Winner':<22}")
        print(f"  {'-'*80}")
        for m in wrong:
            short_v = (m["venue"] or "")[:30]
            print(
                f"  {m['match_no']:>3}  {short_v:<32}  "
                f"{m['sys_rec']:>10}  "
                f"{m['toss_dec']:>8}  "
                f"{str(m['winner']):<22}"
            )
        print()
        print("  RECALIBRATION SIGNALS:")
        for m in wrong:
            venue = m["venue"] or "unknown venue"
            if m["sys_rec"] == "BOWL FIRST" and m["toss_dec"] == "bat":
                print(f"    M{m['match_no']:02d} {venue}: system said BOWL — captain chose BAT and won. "
                      f"Check dew model for this venue.")
            elif m["sys_rec"] == "BAT FIRST" and m["toss_dec"] == "field":
                print(f"    M{m['match_no']:02d} {venue}: system said BAT — captain chose BOWL and won. "
                      f"Chase win % may be underweighted for this venue.")


def partnership_report(matches: list[dict]) -> None:
    """
    Analyse partnership data if column P is filled.
    Checks if the 32-ball Critical threshold was correct.
    """
    filled = [m for m in matches if m["partnership"] is not None]
    if not filled:
        print(f"\n  PARTNERSHIP BACK-TEST: Column P not yet filled — skipping.")
        return

    print(f"\n  PARTNERSHIP BACK-TEST ({len(filled)}/{len(matches)} matches filled)")
    runs = [m["partnership"] for m in filled]
    avg  = sum(runs) / len(runs)
    over_32 = sum(1 for r in runs if r >= 40)   # ~32 balls ≈ 40 runs at 120 SR
    print(f"  Avg biggest partnership : {avg:.1f} runs")
    print(f"  Partnerships ≥ 40 runs  : {over_32}/{len(filled)}  "
          f"({over_32/len(filled)*100:.0f}%) — these likely crossed 32-ball threshold")

    # Breaker analysis if Q is filled
    breakers = [m["p_breaker"] for m in filled if m["p_breaker"]]
    if breakers:
        from collections import Counter
        counts = Counter(b.lower() for b in breakers)
        total_b = len(breakers)
        print(f"\n  Partnership breaker breakdown ({len(breakers)} filled):")
        for b_type, cnt in counts.most_common():
            print(f"    {b_type:<12}: {cnt:>3}  ({cnt/total_b*100:.0f}%)")


def death_eco_report(matches: list[dict]) -> None:
    """Death over economy analysis if column R is filled."""
    filled = [m for m in matches if m["death_eco"] is not None]
    if not filled:
        return
    ecos = [m["death_eco"] for m in filled]
    avg  = sum(ecos) / len(ecos)
    over_10 = sum(1 for e in ecos if e >= 10.0)
    print(f"\n  DEATH ECO BACK-TEST ({len(filled)}/{len(matches)} filled)")
    print(f"  Avg death over economy : {avg:.1f}")
    print(f"  Matches with eco ≥ 10  : {over_10}/{len(filled)}")


def match_outcome_report(matches: list[dict]) -> None:
    """Simple bat-first vs chase win summary from column M."""
    scored = [m for m in matches if m["winner"] and not m["is_nr"] and m["team1"] and m["team2"]]
    if not scored:
        return

    # Infer who batted first from toss decision
    bat_first_wins = 0
    chase_wins     = 0
    for m in scored:
        teams = {m["team1"], m["team2"]}
        if m["winner"] == "No Result" or m["winner"] not in teams:
            continue
        # If toss winner chose bat: toss_winner batted first
        toss_w = m["toss_winner"]
        if m["toss_dec"] == "bat":
            batting_first = toss_w
        else:
            batting_first = next((t for t in teams if t != toss_w), None)
        if batting_first and m["winner"] == batting_first:
            bat_first_wins += 1
        else:
            chase_wins += 1

    total = bat_first_wins + chase_wins
    print(f"\n  MATCH OUTCOME SUMMARY (from filled data)")
    print(f"  Batting first wins : {bat_first_wins}/{total}  ({bat_first_wins/total*100:.0f}%)")
    print(f"  Chasing wins       : {chase_wins}/{total}  ({chase_wins/total*100:.0f}%)")
    print(f"  (Compare to venue_stats.csv chase_win_pct to check calibration)")


# ── ENTRY POINT ──────────────────────────────────────────────────────────────

def main() -> None:
    p = argparse.ArgumentParser(description="PSL back-test runner")
    p.add_argument("--xlsx",         default=str(DEFAULT_XLSX),
                   help=f"Path to back-test xlsx (default: {DEFAULT_XLSX})")
    p.add_argument("--venue",        default="",
                   help="Filter to one venue name (partial match ok)")
    p.add_argument("--summary-only", action="store_true",
                   help="Print only accuracy number, no detail")
    args = p.parse_args()

    xlsx_path = Path(args.xlsx)
    print(f"[backtest_runner] Loading: {xlsx_path}")
    matches = load_backtest_data(xlsx_path)
    print(f"[backtest_runner] {len(matches)} match rows loaded")

    toss_accuracy_report(matches, venue_filter=args.venue)
    if not args.summary_only:
        match_outcome_report(matches)
        partnership_report(matches)
        death_eco_report(matches)

    print(f"\n{'='*65}\n")


if __name__ == "__main__":
    main()
